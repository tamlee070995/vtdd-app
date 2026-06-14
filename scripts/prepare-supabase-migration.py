from __future__ import annotations

import csv
import json
import re
from collections import OrderedDict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
SOURCE_DIR = ROOT / "migration-data"
OUT_DIR = SOURCE_DIR / "out"


FILES = {
    "staff": "Data_Staff.xlsx",
    "products_new": "Data_Moi.xlsx",
    "products_old_phone": "Data_Cu.xlsx",
    "products_old_tablet": "Data_Cu_Tablet.xlsx",
    "pmh_codes": "PMH.xlsx",
    "pincode_requests": "Data_PincodeAudit.xlsx",
    "system_settings": "System_Settings.xlsx",
    "admin_audit": "Admin_Audit.xlsx",
    "quote_logs": "Log_search.xlsx",
}


SCHEMA_SQL = r"""
-- VTDD Supabase schema
-- Chay file nay trong Supabase SQL Editor truoc khi import CSV.
-- Luu y: file nay DROP cac bang cu trong public schema de tao lai dung cau truc.

drop table if exists public.quote_logs cascade;
drop table if exists public.admin_audit cascade;
drop table if exists public.pincode_requests cascade;
drop table if exists public.pmh_codes cascade;
drop table if exists public.products_old cascade;
drop table if exists public.products_new cascade;
drop table if exists public.stores cascade;
drop table if exists public.staff cascade;
drop table if exists public.system_settings cascade;

create table public.staff (
  ma_nv text primary key,
  staff_name text,
  ma_st text,
  store_name text,
  department text,
  password_hash text,
  security_question text,
  security_answer text,
  gmail text,
  status text,
  reset_otp_hash text,
  reset_otp_expires text,
  reset_otp_day text,
  reset_otp_count text,
  need_setup text,
  permission text,
  module_permissions text,
  source_row text,
  imported_at timestamptz not null default now()
);

create table public.stores (
  ma_st text primary key,
  store_name text,
  departments text,
  staff_count text,
  imported_at timestamptz not null default now()
);

create table public.products_new (
  id bigserial primary key,
  brand text,
  product_name text,
  subsidy_ratio text,
  subsidy_ratio_apple text,
  subsidy_amount text,
  category text,
  min_subsidy_amount text,
  min_subsidy_amount_apple text,
  source_row text,
  imported_at timestamptz not null default now()
);

create table public.products_old (
  id bigserial primary key,
  source_sheet text,
  brand text,
  product_name text,
  storage text,
  price_type_1 text,
  price_type_2 text,
  price_type_3 text,
  price_type_4 text,
  price_type_5 text,
  price_type_5_plus text,
  category text,
  mwg_type_1 text,
  mwg_type_2 text,
  source_row text,
  imported_at timestamptz not null default now()
);

create table public.pmh_codes (
  pincode text primary key,
  status text,
  menh_gia text,
  request_id text,
  used_at text,
  used_by text,
  source_row text,
  imported_at timestamptz not null default now()
);

create table public.pincode_requests (
  request_id text primary key,
  created_at_text text,
  ma_st text,
  ma_nv text,
  imei text,
  image_link_1 text,
  image_link_2 text,
  image_link_3 text,
  image_link_4 text,
  image_link_5 text,
  image_link_6 text,
  status text,
  pincode text,
  reject_reason text,
  admin_reviewer text,
  completion_status text,
  menh_gia text,
  old_model text,
  new_model text,
  support_type text,
  source_row text,
  imported_at timestamptz not null default now()
);

create table public.system_settings (
  key text primary key,
  value text,
  type text,
  updated_at_text text,
  updated_by text,
  imported_at timestamptz not null default now()
);

create table public.admin_audit (
  id bigserial primary key,
  time_text text,
  admin text,
  action text,
  target text,
  old_value text,
  new_value text,
  ip text,
  note text,
  source_row text,
  imported_at timestamptz not null default now()
);

create table public.quote_logs (
  id bigserial primary key,
  time_text text,
  action text,
  ma_nv text,
  ma_st text,
  staff_name text,
  flow text,
  product_new text,
  product_old text,
  storage text,
  device_type text,
  old_price text,
  subsidy_brand text,
  subsidy_mwg text,
  customer_total text,
  customer_need_pay text,
  ip text,
  user_agent text,
  source_row text,
  imported_at timestamptz not null default now()
);

create index staff_ma_st_idx on public.staff(ma_st);
create index staff_status_idx on public.staff(status);
create index products_new_category_idx on public.products_new(category);
create index products_old_name_idx on public.products_old(product_name);
create index products_old_category_idx on public.products_old(category);
create index pmh_codes_status_idx on public.pmh_codes(status);
create index pmh_codes_menh_gia_idx on public.pmh_codes(menh_gia);
create index pincode_requests_staff_idx on public.pincode_requests(ma_st, ma_nv);
create index pincode_requests_status_idx on public.pincode_requests(status);
create index quote_logs_staff_idx on public.quote_logs(ma_nv);
create index quote_logs_store_idx on public.quote_logs(ma_st);
"""


def clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return text


def clean_empty_zero(value: Any) -> str:
    text = clean(value)
    return "" if text in {"0", "-", "—"} else text


def clean_permission(value: Any) -> str:
    text = clean(value).lower()
    return text if text in {"admin", "mod"} else ""


def row_has_data(row: Iterable[Any]) -> bool:
    return any(clean(value) for value in row)


def load_rows(file_name: str) -> list[list[Any]]:
    path = SOURCE_DIR / file_name
    if not path.exists():
        raise FileNotFoundError(f"Missing migration file: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def write_csv(name: str, headers: list[str], rows: list[dict[str, str]]) -> None:
    path = OUT_DIR / f"{name}.csv"
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({header: row.get(header, "") for header in headers})


def nonblank_data_rows(rows: list[list[Any]]) -> list[tuple[int, list[Any]]]:
    return [(index, row) for index, row in enumerate(rows[1:], start=2) if row_has_data(row)]


def build_staff() -> tuple[list[dict[str, str]], list[dict[str, str]], dict[str, Any]]:
    headers = [
        "ma_nv",
        "staff_name",
        "ma_st",
        "store_name",
        "department",
        "password_hash",
        "security_question",
        "security_answer",
        "gmail",
        "status",
        "reset_otp_hash",
        "reset_otp_expires",
        "reset_otp_day",
        "reset_otp_count",
        "need_setup",
        "permission",
        "module_permissions",
        "source_row",
    ]
    staff_by_key: OrderedDict[str, dict[str, str]] = OrderedDict()
    duplicate_staff: list[str] = []
    stores: OrderedDict[str, dict[str, Any]] = OrderedDict()

    for source_row, row in nonblank_data_rows(load_rows(FILES["staff"])):
        ma_nv = clean(row[0] if len(row) > 0 else "")
        if not ma_nv:
            continue

        item = {
            "ma_nv": ma_nv,
            "staff_name": clean(row[1] if len(row) > 1 else ""),
            "ma_st": clean(row[2] if len(row) > 2 else ""),
            "store_name": clean(row[3] if len(row) > 3 else ""),
            "department": clean(row[4] if len(row) > 4 else ""),
            "password_hash": clean(row[5] if len(row) > 5 else ""),
            "security_question": clean(row[6] if len(row) > 6 else ""),
            "security_answer": clean(row[7] if len(row) > 7 else ""),
            "gmail": clean(row[8] if len(row) > 8 else ""),
            "status": clean(row[9] if len(row) > 9 else "") or "Standby",
            "reset_otp_hash": clean(row[10] if len(row) > 10 else ""),
            "reset_otp_expires": clean(row[11] if len(row) > 11 else ""),
            "reset_otp_day": clean(row[12] if len(row) > 12 else ""),
            "reset_otp_count": clean(row[13] if len(row) > 13 else "") or "0",
            "need_setup": clean(row[14] if len(row) > 14 else "") or "0",
            "permission": clean_permission(row[15] if len(row) > 15 else ""),
            "module_permissions": clean_empty_zero(row[16] if len(row) > 16 else ""),
            "source_row": str(source_row),
        }

        if ma_nv in staff_by_key:
            duplicate_staff.append(ma_nv)
            continue
        staff_by_key[ma_nv] = item

        ma_st = item["ma_st"]
        if ma_st:
            store = stores.setdefault(
                ma_st,
                {
                    "ma_st": ma_st,
                    "store_name": item["store_name"],
                    "departments": OrderedDict(),
                    "staff_count": 0,
                },
            )
            if item["store_name"] and not store["store_name"]:
                store["store_name"] = item["store_name"]
            if item["department"]:
                store["departments"][item["department"]] = True
            store["staff_count"] += 1

    store_rows = []
    for store in stores.values():
        store_rows.append(
            {
                "ma_st": store["ma_st"],
                "store_name": store["store_name"],
                "departments": " | ".join(store["departments"].keys()),
                "staff_count": str(store["staff_count"]),
            }
        )

    return list(staff_by_key.values()), store_rows, {"headers": headers, "duplicate_staff": duplicate_staff}


def build_products_new() -> list[dict[str, str]]:
    rows = []
    for source_row, row in nonblank_data_rows(load_rows(FILES["products_new"])):
        rows.append(
            {
                "brand": clean(row[0] if len(row) > 0 else ""),
                "product_name": clean(row[1] if len(row) > 1 else ""),
                "subsidy_ratio": clean(row[2] if len(row) > 2 else ""),
                "subsidy_ratio_apple": clean(row[3] if len(row) > 3 else ""),
                "subsidy_amount": clean(row[4] if len(row) > 4 else ""),
                "category": clean(row[5] if len(row) > 5 else ""),
                "min_subsidy_amount": clean(row[6] if len(row) > 6 else ""),
                "min_subsidy_amount_apple": clean(row[7] if len(row) > 7 else ""),
                "source_row": str(source_row),
            }
        )
    return [row for row in rows if row["product_name"]]


def build_products_old(file_name: str, source_sheet: str) -> list[dict[str, str]]:
    rows = []
    for source_row, row in nonblank_data_rows(load_rows(file_name)):
        rows.append(
            {
                "source_sheet": source_sheet,
                "brand": clean(row[0] if len(row) > 0 else ""),
                "product_name": clean(row[1] if len(row) > 1 else ""),
                "storage": clean(row[2] if len(row) > 2 else ""),
                "price_type_1": clean(row[3] if len(row) > 3 else ""),
                "price_type_2": clean(row[4] if len(row) > 4 else ""),
                "price_type_3": clean(row[5] if len(row) > 5 else ""),
                "price_type_4": clean(row[6] if len(row) > 6 else ""),
                "price_type_5": clean(row[7] if len(row) > 7 else ""),
                "price_type_5_plus": clean(row[8] if len(row) > 8 else ""),
                "category": clean(row[9] if len(row) > 9 else ""),
                "mwg_type_1": clean(row[10] if len(row) > 10 else ""),
                "mwg_type_2": clean(row[11] if len(row) > 11 else ""),
                "source_row": str(source_row),
            }
        )
    return [row for row in rows if row["product_name"]]


def build_pmh_codes() -> list[dict[str, str]]:
    rows = []
    seen = set()
    for source_row, row in nonblank_data_rows(load_rows(FILES["pmh_codes"])):
        pincode = clean(row[0] if len(row) > 0 else "")
        if not pincode or pincode in seen:
            continue
        seen.add(pincode)
        rows.append(
            {
                "pincode": pincode,
                "status": clean(row[1] if len(row) > 1 else ""),
                "menh_gia": clean(row[2] if len(row) > 2 else ""),
                "request_id": clean(row[3] if len(row) > 3 else ""),
                "used_at": clean(row[4] if len(row) > 4 else ""),
                "used_by": clean(row[5] if len(row) > 5 else ""),
                "source_row": str(source_row),
            }
        )
    return rows


def build_pincode_requests() -> list[dict[str, str]]:
    rows = []
    for source_row, row in nonblank_data_rows(load_rows(FILES["pincode_requests"])):
        if not (clean(row[0] if len(row) > 0 else "") and clean(row[1] if len(row) > 1 else "")):
            continue
        rows.append(
            {
                "request_id": str(source_row),
                "created_at_text": clean(row[0] if len(row) > 0 else ""),
                "ma_st": clean(row[1] if len(row) > 1 else ""),
                "ma_nv": clean(row[2] if len(row) > 2 else ""),
                "imei": clean(row[3] if len(row) > 3 else ""),
                "image_link_1": clean(row[4] if len(row) > 4 else ""),
                "image_link_2": clean(row[5] if len(row) > 5 else ""),
                "image_link_3": clean(row[6] if len(row) > 6 else ""),
                "image_link_4": clean(row[7] if len(row) > 7 else ""),
                "image_link_5": clean(row[8] if len(row) > 8 else ""),
                "image_link_6": clean(row[9] if len(row) > 9 else ""),
                "status": clean(row[10] if len(row) > 10 else "") or "Pending",
                "pincode": clean(row[11] if len(row) > 11 else ""),
                "reject_reason": clean(row[12] if len(row) > 12 else ""),
                "admin_reviewer": clean(row[13] if len(row) > 13 else ""),
                "completion_status": clean(row[14] if len(row) > 14 else ""),
                "menh_gia": clean(row[15] if len(row) > 15 else ""),
                "old_model": clean(row[16] if len(row) > 16 else ""),
                "new_model": clean(row[17] if len(row) > 17 else ""),
                "support_type": clean(row[18] if len(row) > 18 else "") or "NgoaiDS",
                "source_row": str(source_row),
            }
        )
    return rows


def build_system_settings() -> list[dict[str, str]]:
    rows = []
    seen = set()
    for source_row, row in nonblank_data_rows(load_rows(FILES["system_settings"])):
        key = clean(row[0] if len(row) > 0 else "")
        if not key or key in seen:
            continue
        seen.add(key)
        rows.append(
            {
                "key": key,
                "value": clean(row[1] if len(row) > 1 else ""),
                "type": clean(row[2] if len(row) > 2 else "") or "TEXT",
                "updated_at_text": clean(row[3] if len(row) > 3 else ""),
                "updated_by": clean(row[4] if len(row) > 4 else ""),
            }
        )
    return rows


def build_admin_audit() -> list[dict[str, str]]:
    rows = []
    for source_row, row in nonblank_data_rows(load_rows(FILES["admin_audit"])):
        rows.append(
            {
                "time_text": clean(row[0] if len(row) > 0 else ""),
                "admin": clean(row[1] if len(row) > 1 else ""),
                "action": clean(row[2] if len(row) > 2 else ""),
                "target": clean(row[3] if len(row) > 3 else ""),
                "old_value": clean(row[4] if len(row) > 4 else ""),
                "new_value": clean(row[5] if len(row) > 5 else ""),
                "ip": clean(row[6] if len(row) > 6 else ""),
                "note": clean(row[7] if len(row) > 7 else ""),
                "source_row": str(source_row),
            }
        )
    return [row for row in rows if row["time_text"] or row["action"]]


def build_quote_logs() -> list[dict[str, str]]:
    rows = []
    for source_row, row in nonblank_data_rows(load_rows(FILES["quote_logs"])):
        rows.append(
            {
                "time_text": clean(row[0] if len(row) > 0 else ""),
                "action": clean(row[1] if len(row) > 1 else ""),
                "ma_nv": clean(row[2] if len(row) > 2 else ""),
                "ma_st": clean(row[3] if len(row) > 3 else ""),
                "staff_name": clean(row[4] if len(row) > 4 else ""),
                "flow": clean(row[5] if len(row) > 5 else ""),
                "product_new": clean(row[6] if len(row) > 6 else ""),
                "product_old": clean(row[7] if len(row) > 7 else ""),
                "storage": clean(row[8] if len(row) > 8 else ""),
                "device_type": clean(row[9] if len(row) > 9 else ""),
                "old_price": clean(row[10] if len(row) > 10 else ""),
                "subsidy_brand": clean(row[11] if len(row) > 11 else ""),
                "subsidy_mwg": clean(row[12] if len(row) > 12 else ""),
                "customer_total": clean(row[13] if len(row) > 13 else ""),
                "customer_need_pay": clean(row[14] if len(row) > 14 else ""),
                "ip": clean(row[15] if len(row) > 15 else ""),
                "user_agent": clean(row[16] if len(row) > 16 else ""),
                "source_row": str(source_row),
            }
        )
    return [row for row in rows if row["time_text"] or row["ma_nv"] or row["product_old"]]


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    missing = [name for name in FILES.values() if not (SOURCE_DIR / name).exists()]
    if missing:
        raise SystemExit("Missing files: " + ", ".join(missing))

    staff_rows, store_rows, staff_meta = build_staff()
    products_new = build_products_new()
    products_old = build_products_old(FILES["products_old_phone"], "Data_Cu") + build_products_old(
        FILES["products_old_tablet"], "Data_Cu_Tablet"
    )
    pmh_codes = build_pmh_codes()
    pincode_requests = build_pincode_requests()
    system_settings = build_system_settings()
    admin_audit = build_admin_audit()
    quote_logs = build_quote_logs()

    outputs = {
        "staff": (
            [
                "ma_nv",
                "staff_name",
                "ma_st",
                "store_name",
                "department",
                "password_hash",
                "security_question",
                "security_answer",
                "gmail",
                "status",
                "reset_otp_hash",
                "reset_otp_expires",
                "reset_otp_day",
                "reset_otp_count",
                "need_setup",
                "permission",
                "module_permissions",
                "source_row",
            ],
            staff_rows,
        ),
        "stores": (["ma_st", "store_name", "departments", "staff_count"], store_rows),
        "products_new": (
            [
                "brand",
                "product_name",
                "subsidy_ratio",
                "subsidy_ratio_apple",
                "subsidy_amount",
                "category",
                "min_subsidy_amount",
                "min_subsidy_amount_apple",
                "source_row",
            ],
            products_new,
        ),
        "products_old": (
            [
                "source_sheet",
                "brand",
                "product_name",
                "storage",
                "price_type_1",
                "price_type_2",
                "price_type_3",
                "price_type_4",
                "price_type_5",
                "price_type_5_plus",
                "category",
                "mwg_type_1",
                "mwg_type_2",
                "source_row",
            ],
            products_old,
        ),
        "pmh_codes": (["pincode", "status", "menh_gia", "request_id", "used_at", "used_by", "source_row"], pmh_codes),
        "pincode_requests": (
            [
                "request_id",
                "created_at_text",
                "ma_st",
                "ma_nv",
                "imei",
                "image_link_1",
                "image_link_2",
                "image_link_3",
                "image_link_4",
                "image_link_5",
                "image_link_6",
                "status",
                "pincode",
                "reject_reason",
                "admin_reviewer",
                "completion_status",
                "menh_gia",
                "old_model",
                "new_model",
                "support_type",
                "source_row",
            ],
            pincode_requests,
        ),
        "system_settings": (["key", "value", "type", "updated_at_text", "updated_by"], system_settings),
        "admin_audit": (
            ["time_text", "admin", "action", "target", "old_value", "new_value", "ip", "note", "source_row"],
            admin_audit,
        ),
        "quote_logs": (
            [
                "time_text",
                "action",
                "ma_nv",
                "ma_st",
                "staff_name",
                "flow",
                "product_new",
                "product_old",
                "storage",
                "device_type",
                "old_price",
                "subsidy_brand",
                "subsidy_mwg",
                "customer_total",
                "customer_need_pay",
                "ip",
                "user_agent",
                "source_row",
            ],
            quote_logs,
        ),
    }

    for name, (headers, rows) in outputs.items():
        write_csv(name, headers, rows)

    (OUT_DIR / "schema.sql").write_text(SCHEMA_SQL.strip() + "\n", encoding="utf-8")
    import_order = "\n".join(
        [
            "1. staff.csv -> staff",
            "2. stores.csv -> stores",
            "3. products_new.csv -> products_new",
            "4. products_old.csv -> products_old",
            "5. pmh_codes.csv -> pmh_codes",
            "6. pincode_requests.csv -> pincode_requests",
            "7. system_settings.csv -> system_settings",
            "8. admin_audit.csv -> admin_audit",
            "9. quote_logs.csv -> quote_logs",
        ]
    )
    (OUT_DIR / "import_order.txt").write_text(import_order + "\n", encoding="utf-8")

    report = {
        "source_dir": str(SOURCE_DIR),
        "out_dir": str(OUT_DIR),
        "counts": {name: len(rows) for name, (_, rows) in outputs.items()},
        "staff_duplicate_count": len(staff_meta["duplicate_staff"]),
        "staff_duplicate_samples": staff_meta["duplicate_staff"][:20],
    }
    (OUT_DIR / "report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
